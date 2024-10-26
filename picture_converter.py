from img2table.document import PDF
from img2table.ocr import TesseractOCR
from img2table.document import Image
import os
import fitz
import pandas as pd
from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter
import argparse

def extract_dfs_img(img_name, ocr):
    dfs = []
    image = Image(img_name, detect_rotation=False)
    extracted_tables = image.extract_tables(ocr=ocr,
                                implicit_rows=False,
                                borderless_tables=False,
                                min_confidence=50)  # Экстракция таблиц
    df = extracted_tables[0].df
    df = df.rename(columns=df.iloc[0]).drop(df.index[0])
    dfs.append(df)
    return dfs


def extract_dfs_pdf(pdf_name, ocr):
    pdf_doc = fitz.open(pdf_name)
    pages_num = pdf_doc.page_count
    dfs = []
    for page_num in range(pages_num):
        pdf = PDF(src=pdf_name, pages=[page_num])
        extracted_tables = pdf.extract_tables(ocr=ocr,
                                        implicit_rows=False,
                                        borderless_tables=False,
                                        min_confidence=50)  # Экстракция таблиц
        df = extracted_tables[page_num][0].df
        df = df.rename(columns=df.iloc[0]).drop(df.index[0])
        dfs.append(df)
    return dfs

def rm_extra_columns(dfs):
    for i in range(len(dfs)):
        dfs[i] = dfs[i].loc[:, ~dfs[i].columns.duplicated()]
    return dfs

def unify_headers(dfs):
    headers = dfs[0].columns
    for i in range(len(dfs)):
        dfs[i].columns = headers
        dfs[i] = dfs[i].fillna('')
    return dfs

def fix_date_suffix(text):
    return re.sub(r'(\d{1,2}\.\d{1,2}\.\d{4}|\d{4})\s?[rт]\.?', r'\1 г.', text)

def correct_dates(df):
    for head in df.columns:
        df[head] = df[head].apply(fix_date_suffix)
    return df

def merge_dfs(dfs):
    merged_df = pd.concat(dfs, axis=0)
    merged_df = merged_df.reset_index(drop=True)
    return merged_df
    
    
def merge_cells_row(file_name, merged_df):
    wb = load_workbook(file_name)
    ws = wb.active

    if ws.max_column>1:
        for indx, row in merged_df.iterrows():
            row_data =  pd.Series(row.to_list())
            if len(row_data.unique())==1:
                ws.merge_cells(start_row=indx+2, start_column=1, end_row=indx+2, end_column=ws.max_column)
    wb.save(file_name)
    
def adjust_width_cells(file_name, data):
    wb = load_workbook(file_name)
    ws = wb.active
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
        
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width*30
    
    wb.save(file_name)
    
    

def process_file(file_path):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    # Проверка формата файла
    valid_extensions = ['.jpg', '.jpeg', '.png', '.pdf']
    _, file_extension = os.path.splitext(file_path)
    
    if file_extension.lower() not in valid_extensions:
        raise ValueError(f"Неправильный формат файла: {file_extension}. Допустимые форматы: {', '.join(valid_extensions)}")

    ocr = TesseractOCR(lang='eng+rus')
    
    #получение таблиц
    if file_extension.lower() in '.pdf':
        tables_per_page_df = extract_dfs_pdf(file_path, ocr)
    else:
        tables_per_page_df = extract_dfs_img(file_path, ocr)

    #удаление лишних столбцов 
    tables_per_page_df = rm_extra_columns(tables_per_page_df)
    tables_per_page_df = unify_headers(tables_per_page_df)
    
    #сведение в одну таблицу
    merged_df = merge_dfs(tables_per_page_df)
    
    #постпроцессинг: исправление дат
    merged_df = correct_dates(merged_df)
    
    dir_path = os.path.dirname(file_path)
    table_name = dir_path+'/table.xlsx'
    #сохранение таблицы в excel
    merged_df.to_excel(table_name, index=False)
    
    #форматирование таблицы
    merge_cells_row(table_name, merged_df)
    adjust_width_cells(table_name, merged_df)
    
    return table_name

def main():
    parser = argparse.ArgumentParser(description='Обработка файла.')
    parser.add_argument('file_path', type=str, help='Путь к файлу для обработки')
    args = parser.parse_args()
    
    # Обработка файла
    process_file(args.file_path)
    

if __name__ == '__main__':
    main()